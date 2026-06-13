"""
MediPOS Reports — Views.

Class-based views for generating business reports with optional PDF/Excel
export. All views require authentication. Admin and Pharmacist roles have
full access; Cashier receives a 403 Forbidden redirect to the dashboard.

Export helpers:
- generate_pdf_response(): Creates a styled PDF via ReportLab.
- generate_excel_response(): Creates a styled XLSX via openpyxl.
"""
from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.translation import gettext_lazy as _
from django.views.generic import TemplateView

from apps.inventory.models import Batch
from apps.medicines.models import Category, Medicine
from apps.sales.models import Sale, SaleItem

from .forms import DateRangeForm, ReportFilterForm


# ═══════════════════════════════════════════════════════════════════════════
# Access Control Mixin
# ═══════════════════════════════════════════════════════════════════════════


class StaffAccessMixin(LoginRequiredMixin):
    """
    Restrict access to Admin and Pharmacist roles only.

    Cashier users are redirected to the dashboard with a 403-style
    warning message.  This is enforced in dispatch() so it applies
    to both GET and POST requests.
    """

    def dispatch(self, request, *args, **kwargs):
        """Check role before delegating to the parent handler."""
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        if request.user.role == request.user.Role.CASHIER:
            messages.warning(
                request,
                _('You do not have permission to access reports. '
                  'Please contact an administrator.'),
            )
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)


# ═══════════════════════════════════════════════════════════════════════════
# Export Helpers
# ═══════════════════════════════════════════════════════════════════════════


def generate_pdf_response(title, headers, rows, filename='report.pdf'):
    """
    Generate a styled PDF document using ReportLab and return as an
    HttpResponse with the appropriate content-type.

    Args:
        title (str): Report title displayed at the top.
        headers (list[str]): Column header strings.
        rows (list[list]): List of row lists (each inner list is a data row).
        filename (str): Filename for the Content-Disposition header.

    Returns:
        HttpResponse: PDF file download response.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles['Heading1']
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Build table data — headers + rows
    table_data = [headers] + rows

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, -0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_excel_response(title, headers, rows, filename='report.xlsx'):
    """
    Generate a styled Excel workbook using openpyxl and return as an
    HttpResponse with the appropriate content-type.

    Args:
        title (str): Report title (merged top row).
        headers (list[str]): Column header strings.
        rows (list[list]): List of row lists (each inner list is a data row).
        filename (str): Filename for the Content-Disposition header.

    Returns:
        HttpResponse: XLSX file download response.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Sheet name max 31 chars

    # Title row (merged)
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    col_count = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = header_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Column headers
    col_header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    col_header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[3].height = 20

    # Auto-width columns
    for col_idx in range(1, col_count + 1):
        max_length = 0
        for row_idx in range(1, len(rows) + 3):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(
            max_length + 4, 40
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═══════════════════════════════════════════════════════════════════════════
# Report Views
# ═══════════════════════════════════════════════════════════════════════════


class ReportDashboardView(StaffAccessMixin, TemplateView):
    """
    Landing page listing all available reports as Bootstrap cards.

    Template:
        reports/report_dashboard.html
    """

    template_name = 'reports/report_dashboard.html'


class SalesSummaryView(StaffAccessMixin, TemplateView):
    """
    Sales summary report with date-range filtering.

    GET  → displays the date-range form.
    POST → processes the date range, aggregates Sale data, and renders
           KPI cards, payment-mode breakdown, and daily breakdown.
    """

    template_name = 'reports/sales_summary.html'

    def get_context_data(self, **kwargs):
        """Build context with form and (on POST) aggregated sales data."""
        context = super().get_context_data(**kwargs)
        form = DateRangeForm(
            self.request.POST or None, prefix='sales'
        )
        context['form'] = form

        if self.request.method == 'POST' and form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']

            # Filter completed sales within the date range
            sales_qs = Sale.objects.filter(
                sale_date__date__gte=from_date,
                sale_date__date__lte=to_date,
                status=Sale.Status.COMPLETED,
            )

            # KPI aggregates
            agg = sales_qs.aggregate(
                total_count=Count('id'),
                total_revenue=Sum('grand_total'),
                total_items=Sum('items__quantity'),
            )
            total_count = agg['total_count'] or 0
            total_revenue = agg['total_revenue'] or 0
            total_items = agg['total_items'] or 0
            avg_sale = (total_revenue / total_count) if total_count > 0 else 0

            context.update({
                'from_date': from_date,
                'to_date': to_date,
                'total_count': total_count,
                'total_revenue': total_revenue,
                'total_items': total_items or 0,
                'avg_sale': round(avg_sale, 2),
                'has_results': True,
            })

            # Payment-mode breakdown
            payment_data = (
                sales_qs.values('payment_mode')
                .annotate(count=Count('id'), total=Sum('grand_total'))
                .order_by('-total')
            )
            context['payment_breakdown'] = payment_data

            # Daily sales breakdown
            from django.db.models.functions import TruncDate

            daily_data = (
                sales_qs.annotate(day=TruncDate('sale_date'))
                .values('day')
                .annotate(invoices=Count('id'), amount=Sum('grand_total'))
                .order_by('day')
            )
            context['daily_sales'] = daily_data

            # Export params
            context['export_params'] = (
                f'?from_date={from_date}&to_date={to_date}'
            )

        return context

    def post(self, request, *args, **kwargs):
        """
        Handle POST — also check for export action.

        If 'export' is in POST, delegate to the export handler.
        """
        # Check for export
        export_type = request.POST.get('export')
        if export_type:
            return self._handle_export(request, export_type)
        return super().get(request, *args, **kwargs)

    def _handle_export(self, request, export_type):
        """Generate and return a PDF or Excel export."""
        form = DateRangeForm(request.POST, prefix='sales')
        if not form.is_valid():
            messages.error(request, _('Invalid date range for export.'))
            return redirect('reports:sales_summary')

        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']

        sales_qs = Sale.objects.filter(
            sale_date__date__gte=from_date,
            sale_date__date__lte=to_date,
            status=Sale.Status.COMPLETED,
        )

        from django.db.models.functions import TruncDate

        daily_data = (
            sales_qs.annotate(day=TruncDate('sale_date'))
            .values('day')
            .annotate(invoices=Count('id'), amount=Sum('grand_total'))
            .order_by('day')
        )

        title = f'Sales Summary ({from_date} to {to_date})'
        headers = ['Date', 'Invoices', 'Amount (BDT)']
        rows = [
            [str(d['day']), d['invoices'], float(d['amount'] or 0)]
            for d in daily_data
        ]
        # Totals row
        agg = sales_qs.aggregate(
            total_count=Count('id'), total_revenue=Sum('grand_total')
        )
        rows.append([
            'TOTAL',
            agg['total_count'] or 0,
            float(agg['total_revenue'] or 0),
        ])

        if export_type == 'pdf':
            return generate_pdf_response(
                title, headers, rows, f'sales_summary_{from_date}_{to_date}.pdf'
            )
        else:
            return generate_excel_response(
                title, headers, rows, f'sales_summary_{from_date}_{to_date}.xlsx'
            )


class ProfitLossView(StaffAccessMixin, TemplateView):
    """
    Profit & Loss report — calculates cost vs revenue per medicine.

    Groups SaleItem records by medicine, computes total quantity sold
    and revenue, then multiplies by the medicine's purchase price for
    cost, deriving profit and margin percentage.
    """

    template_name = 'reports/profit_loss.html'

    def get_context_data(self, **kwargs):
        """Build context with form and profit/loss analysis."""
        context = super().get_context_data(**kwargs)
        form = DateRangeForm(self.request.POST or None, prefix='pl')
        context['form'] = form

        if self.request.method == 'POST' and form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']

            # Group SaleItem by medicine for completed sales in range
            items_qs = SaleItem.objects.filter(
                sale__sale_date__date__gte=from_date,
                sale__sale_date__date__lte=to_date,
                sale__status=Sale.Status.COMPLETED,
            ).values(
                'medicine_id',
                'medicine__name',
                'medicine__category__name',
                'medicine__purchase_price',
            ).annotate(
                qty_sold=Sum('quantity'),
                revenue=Sum('line_total'),
            ).order_by('-revenue')

            results = []
            total_qty = 0
            total_revenue = 0
            total_cost = 0

            for item in items_qs:
                qty = item['qty_sold'] or 0
                rev = item['revenue'] or 0
                cost_price = item['medicine__purchase_price'] or 0
                cost = qty * cost_price
                profit = rev - cost
                margin = (profit / rev * 100) if rev > 0 else 0

                results.append({
                    'medicine_name': item['medicine__name'],
                    'category': item['medicine__category__name'] or '-',
                    'qty_sold': qty,
                    'revenue': rev,
                    'cost': cost,
                    'profit': profit,
                    'margin': round(margin, 1),
                })
                total_qty += qty
                total_revenue += rev
                total_cost += cost

            total_profit = total_revenue - total_cost
            total_margin = (
                (total_profit / total_revenue * 100) if total_revenue > 0 else 0
            )

            context.update({
                'from_date': from_date,
                'to_date': to_date,
                'results': results,
                'total_qty': total_qty,
                'total_revenue': total_revenue,
                'total_cost': total_cost,
                'total_profit': total_profit,
                'total_margin': round(total_margin, 1),
                'has_results': True,
            })

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST and optional export."""
        export_type = request.POST.get('export')
        if export_type:
            return self._handle_export(request, export_type)
        return super().get(request, *args, **kwargs)

    def _handle_export(self, request, export_type):
        """Export P&L data to PDF or Excel."""
        form = DateRangeForm(request.POST, prefix='pl')
        if not form.is_valid():
            messages.error(request, _('Invalid date range.'))
            return redirect('reports:profit_loss')

        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']

        items_qs = SaleItem.objects.filter(
            sale__sale_date__date__gte=from_date,
            sale__sale_date__date__lte=to_date,
            sale__status=Sale.Status.COMPLETED,
        ).values(
            'medicine_id',
            'medicine__name',
            'medicine__purchase_price',
        ).annotate(
            qty_sold=Sum('quantity'),
            revenue=Sum('line_total'),
        ).order_by('-revenue')

        title = f'Profit & Loss ({from_date} to {to_date})'
        headers = ['Medicine', 'Qty Sold', 'Revenue', 'Cost', 'Profit', 'Margin %']
        rows = []

        for item in items_qs:
            qty = item['qty_sold'] or 0
            rev = float(item['revenue'] or 0)
            cost = qty * float(item['medicine__purchase_price'] or 0)
            profit = rev - cost
            margin = (profit / rev * 100) if rev > 0 else 0
            rows.append([
                item['medicine__name'], qty, rev, cost,
                profit, round(margin, 1),
            ])

        if export_type == 'pdf':
            return generate_pdf_response(
                title, headers, rows, f'profit_loss_{from_date}_{to_date}.pdf'
            )
        else:
            return generate_excel_response(
                title, headers, rows, f'profit_loss_{from_date}_{to_date}.xlsx'
            )


class ProductSalesView(StaffAccessMixin, TemplateView):
    """
    Product sales breakdown — shows sales grouped by medicine with
    optional category and medicine filters.
    """

    template_name = 'reports/product_sales.html'

    def get_context_data(self, **kwargs):
        """Build context with form and product-sales breakdown."""
        context = super().get_context_data(**kwargs)
        form = ReportFilterForm(self.request.POST or None, prefix='ps')
        context['form'] = form

        if self.request.method == 'POST' and form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']
            category = form.cleaned_data.get('category')
            medicine = form.cleaned_data.get('medicine')

            items_qs = SaleItem.objects.filter(
                sale__sale_date__date__gte=from_date,
                sale__sale_date__date__lte=to_date,
                sale__status=Sale.Status.COMPLETED,
            )

            if category:
                items_qs = items_qs.filter(medicine__category=category)
            if medicine:
                items_qs = items_qs.filter(medicine=medicine)

            results = (
                items_qs.values(
                    'medicine_id',
                    'medicine__name',
                    'medicine__category__name',
                    'medicine__purchase_price',
                )
                .annotate(
                    qty_sold=Sum('quantity'),
                    revenue=Sum('line_total'),
                )
                .order_by('-qty_sold')
            )

            enriched = []
            total_qty = 0
            total_revenue = 0
            total_profit = 0

            for r in results:
                qty = r['qty_sold'] or 0
                rev = r['revenue'] or 0
                cost = qty * (r['medicine__purchase_price'] or 0)
                profit = rev - cost
                enriched.append({
                    'medicine_name': r['medicine__name'],
                    'category': r['medicine__category__name'] or '-',
                    'qty_sold': qty,
                    'revenue': rev,
                    'profit': profit,
                })
                total_qty += qty
                total_revenue += rev
                total_profit += profit

            context.update({
                'from_date': from_date,
                'to_date': to_date,
                'results': enriched,
                'total_qty': total_qty,
                'total_revenue': total_revenue,
                'total_profit': total_profit,
                'has_results': True,
            })

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST and optional export."""
        export_type = request.POST.get('export')
        if export_type:
            return self._handle_export(request, export_type)
        return super().get(request, *args, **kwargs)

    def _handle_export(self, request, export_type):
        """Export product sales to PDF or Excel."""
        form = ReportFilterForm(request.POST, prefix='ps')
        if not form.is_valid():
            messages.error(request, _('Invalid filter parameters.'))
            return redirect('reports:product_sales')

        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']
        category = form.cleaned_data.get('category')
        medicine = form.cleaned_data.get('medicine')

        items_qs = SaleItem.objects.filter(
            sale__sale_date__date__gte=from_date,
            sale__sale_date__date__lte=to_date,
            sale__status=Sale.Status.COMPLETED,
        )
        if category:
            items_qs = items_qs.filter(medicine__category=category)
        if medicine:
            items_qs = items_qs.filter(medicine=medicine)

        results = (
            items_qs.values('medicine__name', 'medicine__category__name')
            .annotate(qty_sold=Sum('quantity'), revenue=Sum('line_total'))
            .order_by('-qty_sold')
        )

        title = f'Product Sales ({from_date} to {to_date})'
        headers = ['Medicine', 'Category', 'Qty Sold', 'Revenue']
        rows = [
            [r['medicine__name'], r['medicine__category__name'] or '-',
             r['qty_sold'] or 0, float(r['revenue'] or 0)]
            for r in results
        ]

        if export_type == 'pdf':
            return generate_pdf_response(
                title, headers, rows, f'product_sales_{from_date}_{to_date}.pdf'
            )
        else:
            return generate_excel_response(
                title, headers, rows, f'product_sales_{from_date}_{to_date}.xlsx'
            )


class StockValuationView(StaffAccessMixin, TemplateView):
    """
    Current stock valuation — shows all active medicines with stock > 0,
    their purchase price, and computed total value.
    """

    template_name = 'reports/stock_valuation.html'

    def get_context_data(self, **kwargs):
        """Build context with current stock valuation data."""
        context = super().get_context_data(**kwargs)

        medicines = Medicine.objects.filter(
            is_active=True,
            stock_quantity__gt=0,
        ).select_related('category').order_by('category__name', 'name')

        results = []
        total_value = 0
        total_qty = 0

        for med in medicines:
            value = med.stock_quantity * med.purchase_price
            results.append({
                'name': med.name,
                'brand': med.brand or '',
                'category': med.category.name if med.category else '-',
                'stock_qty': med.stock_quantity,
                'purchase_price': med.purchase_price,
                'total_value': value,
            })
            total_value += value
            total_qty += med.stock_quantity

        context.update({
            'results': results,
            'total_value': total_value,
            'total_qty': total_qty,
            'medicine_count': len(results),
        })
        return context

    def post(self, request, *args, **kwargs):
        """Handle POST for export."""
        export_type = request.POST.get('export')
        if export_type:
            return self._handle_export(request, export_type)
        return super().get(request, *args, **kwargs)

    def _handle_export(self, request, export_type):
        """Export stock valuation to PDF or Excel."""
        medicines = Medicine.objects.filter(
            is_active=True, stock_quantity__gt=0,
        ).select_related('category').order_by('category__name', 'name')

        title = 'Stock Valuation Report'
        headers = ['Medicine', 'Category', 'Stock Qty', 'Purchase Price', 'Total Value']
        rows = []
        for med in medicines:
            rows.append([
                med.name,
                med.category.name if med.category else '-',
                med.stock_quantity,
                float(med.purchase_price),
                float(med.stock_quantity * med.purchase_price),
            ])

        if export_type == 'pdf':
            return generate_pdf_response(title, headers, rows, 'stock_valuation.pdf')
        else:
            return generate_excel_response(title, headers, rows, 'stock_valuation.xlsx')


class ExpiryReportView(StaffAccessMixin, TemplateView):
    """
    Batch expiry report with status filter (ALL, EXPIRED, EXPIRING_30,
    EXPIRING_60, OK).
    """

    template_name = 'reports/expiry_report.html'

    EXPIRY_STATUSES = [
        ('ALL', 'All Batches'),
        ('EXPIRED', 'Expired'),
        ('EXPIRING_30', 'Expiring Within 30 Days'),
        ('EXPIRING_60', 'Expiring Within 60 Days'),
        ('OK', 'Safe (60+ Days)'),
    ]

    def get_context_data(self, **kwargs):
        """Build context with expiry analysis."""
        context = super().get_context_data(**kwargs)
        status_filter = self.request.GET.get('status', 'ALL')
        context['status_filter'] = status_filter
        context['expiry_statuses'] = self.EXPIRY_STATUSES

        today = date.today()
        cutoff_30 = today + timedelta(days=30)
        cutoff_60 = today + timedelta(days=60)

        batches_qs = Batch.objects.filter(
            quantity__gt=0, is_active=True
        ).select_related('medicine', 'supplier').order_by('expiry_date')

        # Apply status filter
        if status_filter == 'EXPIRED':
            batches_qs = batches_qs.filter(expiry_date__lt=today)
        elif status_filter == 'EXPIRING_30':
            batches_qs = batches_qs.filter(
                expiry_date__gte=today, expiry_date__lte=cutoff_30
            )
        elif status_filter == 'EXPIRING_60':
            batches_qs = batches_qs.filter(
                expiry_date__gte=today, expiry_date__lte=cutoff_60
            )
        elif status_filter == 'OK':
            batches_qs = batches_qs.filter(expiry_date__gt=cutoff_60)

        # Summary counts
        all_batches = Batch.objects.filter(quantity__gt=0, is_active=True)
        context['count_total'] = all_batches.count()
        context['count_expired'] = all_batches.filter(expiry_date__lt=today).count()
        context['count_expiring_30'] = all_batches.filter(
            expiry_date__gte=today, expiry_date__lte=cutoff_30,
        ).count()
        context['count_expiring_60'] = all_batches.filter(
            expiry_date__gte=today, expiry_date__lte=cutoff_60,
        ).count()
        context['count_ok'] = all_batches.filter(expiry_date__gt=cutoff_60).count()

        # Results
        results = []
        for batch in batches_qs:
            days = batch.days_until_expiry
            if days < 0:
                badge = 'danger'
                status_text = 'Expired'
            elif days <= 30:
                badge = 'danger'
                status_text = 'Critical'
            elif days <= 60:
                badge = 'warning'
                status_text = 'Warning'
            else:
                badge = 'success'
                status_text = 'OK'

            results.append({
                'medicine_name': batch.medicine.name,
                'batch_no': batch.batch_no,
                'supplier_name': batch.supplier.name if batch.supplier else '-',
                'quantity': batch.quantity,
                'expiry_date': batch.expiry_date,
                'days': days,
                'badge': badge,
                'status_text': status_text,
            })

        context['results'] = results
        return context

    def post(self, request, *args, **kwargs):
        """Handle POST for export."""
        export_type = request.POST.get('export')
        status_filter = request.POST.get('status', 'ALL')
        if export_type:
            return self._handle_export(request, export_type, status_filter)
        return self.get(request, *args, **kwargs)

    def _handle_export(self, request, export_type, status_filter):
        """Export expiry report."""
        today = date.today()
        cutoff_30 = today + timedelta(days=30)
        cutoff_60 = today + timedelta(days=60)

        batches_qs = Batch.objects.filter(
            quantity__gt=0, is_active=True,
        ).select_related('medicine', 'supplier').order_by('expiry_date')

        if status_filter == 'EXPIRED':
            batches_qs = batches_qs.filter(expiry_date__lt=today)
        elif status_filter == 'EXPIRING_30':
            batches_qs = batches_qs.filter(
                expiry_date__gte=today, expiry_date__lte=cutoff_30
            )
        elif status_filter == 'EXPIRING_60':
            batches_qs = batches_qs.filter(
                expiry_date__gte=today, expiry_date__lte=cutoff_60
            )
        elif status_filter == 'OK':
            batches_qs = batches_qs.filter(expiry_date__gt=cutoff_60)

        title = 'Expiry Report'
        headers = ['Medicine', 'Batch No', 'Supplier', 'Qty', 'Expiry Date', 'Days', 'Status']
        rows = []
        for batch in batches_qs:
            days = batch.days_until_expiry
            rows.append([
                batch.medicine.name, batch.batch_no,
                batch.supplier.name if batch.supplier else '-',
                batch.quantity, str(batch.expiry_date), days,
                'Expired' if days < 0 else ('Critical' if days <= 30 else (
                    'Warning' if days <= 60 else 'OK')),
            ])

        if export_type == 'pdf':
            return generate_pdf_response(title, headers, rows, 'expiry_report.pdf')
        else:
            return generate_excel_response(title, headers, rows, 'expiry_report.xlsx')


class TopSellingView(StaffAccessMixin, TemplateView):
    """
    Top-selling medicines report — shows the N most-sold medicines
    within a date range.
    """

    template_name = 'reports/top_selling.html'

    LIMIT_CHOICES = [('10', 'Top 10'), ('25', 'Top 25'), ('50', 'Top 50')]

    def get_context_data(self, **kwargs):
        """Build context with form and ranked results."""
        context = super().get_context_data(**kwargs)
        form = DateRangeForm(self.request.POST or None, prefix='top')
        context['form'] = form
        context['limit_choices'] = self.LIMIT_CHOICES

        limit = int(self.request.POST.get('limit', 10) if self.request.method == 'POST' else 10)
        context['selected_limit'] = str(limit)

        if self.request.method == 'POST' and form.is_valid():
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']

            results = (
                SaleItem.objects.filter(
                    sale__sale_date__date__gte=from_date,
                    sale__sale_date__date__lte=to_date,
                    sale__status=Sale.Status.COMPLETED,
                )
                .values(
                    'medicine_id',
                    'medicine__name',
                    'medicine__category__name',
                )
                .annotate(
                    qty_sold=Sum('quantity'),
                    revenue=Sum('line_total'),
                )
                .order_by('-qty_sold')[:limit]
            )

            ranked = []
            for rank, r in enumerate(results, 1):
                ranked.append({
                    'rank': rank,
                    'medicine_name': r['medicine__name'],
                    'category': r['medicine__category__name'] or '-',
                    'qty_sold': r['qty_sold'] or 0,
                    'revenue': r['revenue'] or 0,
                })

            context.update({
                'from_date': from_date,
                'to_date': to_date,
                'results': ranked,
                'has_results': True,
            })

        return context

    def post(self, request, *args, **kwargs):
        """Handle POST and optional export."""
        export_type = request.POST.get('export')
        if export_type:
            return self._handle_export(request, export_type)
        return super().get(request, *args, **kwargs)

    def _handle_export(self, request, export_type):
        """Export top-selling data."""
        form = DateRangeForm(request.POST, prefix='top')
        if not form.is_valid():
            messages.error(request, _('Invalid date range.'))
            return redirect('reports:top_selling')

        from_date = form.cleaned_data['from_date']
        to_date = form.cleaned_data['to_date']
        limit = int(request.POST.get('limit', 10))

        results = (
            SaleItem.objects.filter(
                sale__sale_date__date__gte=from_date,
                sale__sale_date__date__lte=to_date,
                sale__status=Sale.Status.COMPLETED,
            )
            .values('medicine__name', 'medicine__category__name')
            .annotate(qty_sold=Sum('quantity'), revenue=Sum('line_total'))
            .order_by('-qty_sold')[:limit]
        )

        title = f'Top {limit} Selling Medicines ({from_date} to {to_date})'
        headers = ['Rank', 'Medicine', 'Category', 'Qty Sold', 'Revenue']
        rows = []
        for rank, r in enumerate(results, 1):
            rows.append([
                rank, r['medicine__name'],
                r['medicine__category__name'] or '-',
                r['qty_sold'] or 0,
                float(r['revenue'] or 0),
            ])

        if export_type == 'pdf':
            return generate_pdf_response(
                title, headers, rows, f'top_selling_{from_date}_{to_date}.pdf'
            )
        else:
            return generate_excel_response(
                title, headers, rows, f'top_selling_{from_date}_{to_date}.xlsx'
            )